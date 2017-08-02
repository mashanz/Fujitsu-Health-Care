import os
import openpyxl
#=====================================================================
# READ EXCELS FILE USING EXTERNAL PYTHON LIBRARY
#=====================================================================
print("[database] loading")
from openpyxl import load_workbook
wb = load_workbook(filename='13.xlsx', read_only=True)
ws = wb['Sheet1']

#=====================================================================
# CONVERT EXCEL TO ARRAYS
#=====================================================================
print("[database] converting")
def iter_rows(ws):
 for row in ws.iter_rows():
  yield [cell.value for cell in row]
tabel = list(iter_rows(ws))

#=====================================================================
# REFORMATING ARRAYS - INFORMATION ROWS FINDER
#=====================================================================
c_utc_JST=0
c_steps=0
c_pulse=0
c_calory=0
c_transType=0
def find_row(tabel):
 count=0
 print("[database] finding necesseries data")
 while(1):
  if (count==len(tabel[0])):
   global c_transType
   global c_steps
   global c_pulse
   global c_calory
   global c_utc_JST
   break
  if (tabel[0][count]=="steps"):
   c_steps=count
  if (tabel[0][count]=="pulse"):
   c_pulse=count
  if (tabel[0][count]=="calory"):
   c_calory=count
  if (tabel[0][count]=="utc_JST"):
   c_utc_JST=count
  if (tabel[0][count]=="transType"):
   c_transType=count
  count+=1
find_row(tabel)

#=====================================================================
# KIND OF INPUT DATA START PARAMETERS
#=====================================================================
name="Hans (Hanjara Cahya Adhyatma)"
gender="male"
age=22
weigh=60
heigh=170
pulse=0
avcnt=0
i=0
elevator=0
walking=0
stay=0
stair=0
stp=0
clr=0.0
n=0
person=[]
database=[]
all_pulse=[]
all_stair=[]
all_elevator=[]
all_walking=[]
all_stp=[]
all_stay=[]
all_clr=[]
#=====================================================================
# INFORMATION FILTERING
#=====================================================================
print("[database] collecting necesseries data")
for row in ws.rows:
 if(tabel[i][c_transType]=="7(エレベータ降り中)") or \
  (tabel[i][c_transType]=="6(エレベータ昇り中)"):
  all_elevator.extend([[tabel[i][c_utc_JST], tabel[i][c_transType]]])
  elevator+=1
 if(tabel[i][c_transType]=="2(歩行中)"):
  all_walking.extend([[tabel[i][c_utc_JST], tabel[i][c_transType]]])
  walking+=1
 if(tabel[i][c_transType]=="1(静止中)"):
  all_stay.extend([[tabel[i][c_utc_JST], tabel[i][c_transType]]])
  stay+=1
 if(tabel[i][c_transType]=="5(階段降り中)"):
  all_stair.extend([[tabel[i][c_utc_JST], tabel[i][c_transType]]])
  stair+=1
 if(tabel[i][c_steps] != 0) and \
  (tabel[i][c_steps] != None) and \
  (tabel[i][c_steps] != "steps"):
  all_stp.extend([[tabel[i][c_utc_JST], tabel[i][c_steps]]])
  stp+=int(tabel[i][c_steps])
 if(tabel[i][c_calory] != 0) and \
  (tabel[i][c_calory] != None) and \
  (tabel[i][c_calory] != "calory") and \
  (tabel[i][c_calory] != ""):
  all_clr.extend([[tabel[i][c_utc_JST], tabel[i][c_calory]]])
  clr+=float(tabel[i][c_calory])
 if(tabel[i][c_pulse] != 0) and \
  (tabel[i][c_pulse] != None) and \
  (tabel[i][c_pulse] != "pulse"):
  all_pulse.extend([[tabel[i][c_utc_JST], tabel[i][c_pulse]]])
  pulse+=int(tabel[i][c_pulse])
  avcnt+=1
 i+=1 
 
#=====================================================================
# INFORMATION PRINTING
#=====================================================================

print()
print("Name \t",name)
print("Gender \t",gender)
print(age, "\t years old")
print(weigh, "\t kg")
print(heigh, "\t cm")
print(round(pulse/avcnt, 2), "\t bpm")
print(stair,"\t times detected using stairs")
print(elevator,"\t times detected using elevator")
print(walking, "\t times detected walking")
print(stp,"\t total steps")
print(stay, "\t times detected stay")
print(round(clr, 2),"\t kcal burned total?")
"""
global name
global gender
global age
global weigh
global heigh
global pulse
global avcnt
global stair
global elevator
global walking
global stp
global stay
global clr
"""
#=====================================================================
# new NECESSERIES　database
#=====================================================================
print("[database] save collected data")
person=[name,gender,age,weigh,heigh,round(pulse/avcnt, 2),stair,elevator,walking,stp,stay,round(clr, 2)]
database=[person,gender,age,weigh,heigh,all_pulse,all_stair,all_elevator,all_walking,all_stp,all_stay,all_clr]
"""
it=0
while(1):
 if (it==len(database[9])):
  break
 print(database[9][it])
 it+=1
"""
print("[database] finish")